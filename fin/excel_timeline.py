#encoding=utf-8
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from workalendar.asia import SouthKorea
import datetime
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side


import pathlib
import importlib
import sys


from PySide6 import QtWidgets, QtCore
from ui import schedule_list_ui
from libs.qt import library as qt_lib

importlib.reload(schedule_list_ui)
importlib.reload(qt_lib)


class Signals(QtCore.QObject):
    save_complete = QtCore.Signal(str)


class ProjectSchedule:
    def __init__(
        self,
        schedule_end_path: str,
        in_date: datetime.datetime,
        out_date: datetime.datetime,
        project_name: str,
        project: str,
        schedule_list: str,
    ):
        super().__init__()
        self.signals = Signals()
        self.__schedule_end_path = schedule_end_path
        self.__schedule_in_date = in_date
        self.__schedule_out_date = out_date
        self.__project_name = project_name
        self.__project = project
        self.listWidget_schedule = schedule_list
        self.date_range = [
            self.__schedule_in_date + datetime.timedelta(days=x)
            for x in range((self.__schedule_out_date - self.__schedule_in_date).days + 1)
        ]

        self.holidays = self.get_holidays(self.__schedule_in_date.year, self.__schedule_out_date.year)
        self.wb = Workbook()
        self.ws = self.wb.active
        self.__is_stop = False

    @property
    def get_project_name(self):
        return self.__project_name

    @property
    def is_stop_s(self):
        return self.__is_stop

    @is_stop_s.setter
    def is_stop_s(self, flag: bool):
        self.__is_stop = flag

    def get_holidays(self, start_year, end_year):
        holidays = []
        kr_calendar = SouthKorea()  # SouthKorea 인스턴스 생성
        for year in range(start_year, end_year + 1):
            year_holidays = kr_calendar.holidays(year)  # 해당 연도의 공휴일 목록을 가져옴
            for holiday in year_holidays:
                holidays.append(holiday[0])  # 공휴일 날짜만 holidays 리스트에 추가
        return holidays

    def add_project_to_list(self, project_name):
        item = QtWidgets.QListWidgetItem(project_name)
        self.listWidget_schedule.addItem(item)

    def create_workbook(self):
        # 스타일 정의
        red_font = Font(color="FF0000", bold=True)
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        pastel_colors = ["FFD7A9", "D4A5A5", "FFABAB", "ABE09F", "A6C7E1", "E5B0E1"]

        row = 3  # 프로젝트 이름이 있는 행부터 시작
        for project in self.__project:
            cell = self.ws.cell(row=row, column=1, value=project)
            cell.fill = PatternFill(
                start_color=pastel_colors[(row - 2) % len(pastel_colors)],
                end_color=pastel_colors[(row - 2) % len(pastel_colors)],
                fill_type="solid",
            )
            if isinstance(cell.value, datetime.datetime):  # 날짜인 경우에만 주말 또는 공휴일 여부 확인
                if cell.value.weekday() >= 5 or cell.value in self.holidays:
                    cell.font = red_font  # 주말 또는 공휴일인 경우 빨간색 폰트 적용

            row += 1

        # 월별 날짜 범위 및 첫 번째 행에 월 병합
        month_col = {}
        for date in self.date_range:
            month = date.strftime("%Y-%m")
            if month not in month_col:
                month_col[month] = []
            month_col[month].append(date)

        # 월별로 날짜 범위 및 첫 번째 행에 월 병합
        col = 2
        for month, days in month_col.items():
            merge_start = col
            for day in days:
                self.ws.cell(row=1, column=col, value=day.strftime("%Y-%m-%d"))
                is_holiday = day.weekday() >= 5 or day in self.holidays
                if is_holiday:
                    for row in range(3, len(self.__project) + 3):  # 프로젝트 이름이 있는 행까지 채워줌
                        self.ws.cell(row=row, column=col).fill = gray_fill
                col += 1
            # 월별로 첫 번째 행 병합
            self.ws.merge_cells(start_row=1, start_column=merge_start, end_row=1, end_column=col - 1)
            self.ws.cell(row=1, column=merge_start, value=month).alignment = Alignment(horizontal="center")

        # 프로젝트 이름 표시
        row = 3  # 프로젝트 이름이 있는 행부터 시작
        for project in self.__project:
            self.ws.cell(row=row, column=1, value=project).fill = PatternFill(
                start_color=pastel_colors[(row - 2) % len(pastel_colors)],
                end_color=pastel_colors[(row - 2) % len(pastel_colors)],
                fill_type="solid",
            )
            self.ws.cell(row=row, column=1).font = Font(bold=True)  # 프로젝트 이름을 볼드체로 표시
            row += 1

        # 날짜 범위 표시
        for col, date in enumerate(self.date_range, start=2):  # 첫 번째 행은 프로젝트 이름이 아닌 부분이므로 2부터 시작
            cell = self.ws.cell(row=2, column=col, value=date.strftime("%Y-%m-%d"))
            cell.alignment = Alignment(horizontal="center")
            if date.weekday() >= 5 or date in self.holidays:
                cell.font = Font(color="FF0000")
                cell.fill = gray_fill

        self.ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        self.ws["A1"].value = "project_name"
        self.ws["A1"].font = Font(bold=True)

        for row in range(1, 3):
            for col, date in enumerate(self.date_range, start=1):
                self.ws.cell(row=row, column=col).border = Border(
                    top=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="thin")
                )

        # 프로젝트 행까지 선 추가
        for row in range(2, len(self.__project) + 3):
            for col, date in enumerate(self.date_range, start=1):
                cell = self.ws.cell(row=row, column=col)
                cell.border = Border(top=Side(style="thin"), bottom=Side(style="thin"))
            self.ws.cell(row=row, column=len(self.date_range) + 1).border = Border(
                right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
            )
        # 각 열에 대해 상단 및 하단 선 추가
        for col in range(1, len(self.date_range) + 2):
            for row in range(1, len(self.__project) + 3):
                cell = self.ws.cell(row=row, column=col)
                cell.border = Border(
                    top=Side(style="thin"), bottom=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin")
                )

        filename = "project_schedule.xlsx"
        save_path = pathlib.Path(self.__schedule_end_path) / filename
        self.wb.save(save_path)
        print(f"엑셀 파일이 '{save_path}' 이름으로 저장되었습니다.")


class ExcelSchedule(QtWidgets.QMainWindow, schedule_list_ui.Ui_Form):
    def __init__(self):
        super(ExcelSchedule, self).__init__()
        self.setupUi(self)
        self.signals = Signals()
        self.setWindowFlags(self.windowFlags() | QtCore.Qt.WindowStaysOnTopHint)
        self.__worker = None
        self.textEdit_schedule_name_input.text()
        self.textEdit_schedule_path.text().strip()
        self.dateEdit_shedule_in_date.date().toString("yyyy-MM-dd").strip()
        self.dateEdit_shedule_out_date.date().toString("yyyy-MM-dd").strip()

        self.toolButton__schedule_end_path.setText("end_path")
        self.toolButton__schedule_end_path.clicked.connect(self.slot_schedule_dir)

        self.pushButton_schedule_plus.clicked.connect(self.add_project_to_list)
        self.pushButton_schedule_minus.clicked.connect(self.remove_project)
        self.pushButton_schedule_generate.clicked.connect(self.schedule_toggle_start_stop)
        self.textEdit_schedule_name_input.mousePressEvent = self.clear_initial_text
        self.textEdit_schedule_name_input.setReadOnly(True)

    def clear_initial_text(self, event):
        if self.textEdit_schedule_name_input.isReadOnly():
            self.textEdit_schedule_name_input.setReadOnly(False)
            self.textEdit_schedule_name_input.clear()
        super().mousePressEvent(event)

    def schedule_toggle_start_stop(self):
        if not self.__worker:
            self.initialize_worker()
        else:
            self.__worker = None  # 작업 중지 로직
            self.pushButton_schedule_generate.setText("Start")

    def initialize_worker(self):
        # 필요한 인자 값들을 설정
        schedule_end_path = self.textEdit_schedule_path.text().strip()
        in_date = datetime.datetime.strptime(self.dateEdit_shedule_in_date.date().toString("yyyy-MM-dd"), "%Y-%m-%d")
        out_date = datetime.datetime.strptime(self.dateEdit_shedule_out_date.date().toString("yyyy-MM-dd"), "%Y-%m-%d")
        project_name = self.textEdit_schedule_name_input.text()
        project_count = self.listWidget_schedule.count()  # Get total project count
        project = [self.listWidget_schedule.item(i).text() for i in range(project_count)]
        self.__worker = ProjectSchedule(
            schedule_end_path, in_date, out_date, project_name, project, self.listWidget_schedule
        )
        self.__worker.signals.save_complete.connect(self.show_save_complete_message)
        self.__worker.create_workbook()
        self.pushButton_schedule_generate.setText("Stop")

    def show_save_complete_message(self, save_path):
        QtWidgets.QMessageBox.information(self, "저장 완료", f"파일이 성공적으로 저장되었습니다:\n{save_path}")
        self.__worker = None  # 작업이 완료된 후 worker 인스턴스 초기화
        self.pushButton_schedule_generate.setText("Start")

    def slot_schedule_dir(self):
        sel_dir = QtWidgets.QFileDialog.getExistingDirectory(self, "파일을 선택하세요!")
        if sel_dir:
            self.textEdit_schedule_path.setText(sel_dir)

    def add_project_to_list(self):
        project_name = self.textEdit_schedule_name_input.text()
        if project_name:
            self.listWidget_schedule.addItem(project_name)
            self.listWidget_schedule.scrollToBottom()

    def remove_project(self):
        selected_item = self.listWidget_schedule.currentItem()
        if selected_item:
            row = self.listWidget_schedule.row(selected_item)
            self.listWidget_schedule.takeItem(row)

    def slot_schedule_dir(self):
        sel_dir = QtWidgets.QFileDialog.getExistingDirectory(self, "파일을 선택하세요!")
        if sel_dir:
            self.textEdit_schedule_path.setText(sel_dir)  #

    def create_workbook(self):
        self.__worker.create_workbook()
        filename = f"schedule.xlsx"
        save_path = pathlib.Path(self.__worker._ProjectSchedule__schedule_end_path) / filename
        self.__worker.wb.save(save_path)


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    sch = ExcelSchedule()
    sch.show()
    sys.exit(app.exec())
