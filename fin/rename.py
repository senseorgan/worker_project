import sys
import importlib
import os
import pathlib
import pandas as pd
from openpyxl import load_workbook
from PySide6 import QtWidgets, QtCore
from ui import rename_ui
from libs.qt import library as qt_lib
from libs.system import sys_library as sys_lib

importlib.reload(rename_ui)
importlib.reload(qt_lib)
importlib.reload(sys_lib)


class MultipleData:
    def __init__(self, data: dict):
        self.data = data


class Signals(QtCore.QObject):
    sig_update = QtCore.Signal(int)
    sig_info = QtCore.Signal(str, str)
    sig_done = QtCore.Signal()


class WorkThread(QtCore.QThread):
    def __init__(self, file_path=pathlib.Path, excel_path=pathlib.Path):
        super().__init__()
        self.signals = Signals()
        self.__is_stop = False
        self.lineEdit__file_path = file_path
        self.lineEdit__excel_path = excel_path
        self.start_number = 1000

    def get_new_name(self):
        df = pd.read_excel(self.lineEdit__excel_path, header=None)
        rename_map = dict(zip(df.iloc[:, 0], df.iloc[:, 1]))
        print(rename_map)
        return rename_map

    def set_stop(self, flag):
        self.__is_stop = flag

    def run(self):
        rename_map = self.get_new_name()

        for root, dirs, files in os.walk(self.lineEdit__file_path, topdown=False):
            local_start_number = self.start_number  # 각 폴더마다 시작 번호를 초기화
            for dir_name in dirs:
                original_dir_path = os.path.join(root, dir_name)
                if dir_name in rename_map:
                    new_dir_name = rename_map[dir_name]
                    new_dir_path = os.path.join(root, new_dir_name)
                    os.rename(original_dir_path, new_dir_path)
                    self.signals.sig_info.emit(original_dir_path, new_dir_path)

            for file in sorted(files):
                if file == ".DS_Store":
                    continue
                original_file_path = os.path.join(root, file)
                file_base_name = file.split(".")[0]  # 확장자 제외한 파일 기본 이름
                if file_base_name in rename_map:
                    new_base_name = rename_map[file_base_name]
                    file_extension = os.path.splitext(file)[1]

                    new_file_name = f"{new_base_name}.{local_start_number}{file_extension}"

                    new_file_path = os.path.join(root, new_file_name)
                    os.rename(original_file_path, new_file_path)
                    local_start_number += 1  # 파일이 성공적으로 리네임될 때마다 번호를 증가
                    self.signals.sig_info.emit(original_file_path, new_file_path)
        self.signals.sig_done.emit()

    def count_total_files(self, folder_path):
        file_count = 0
        for root, subdirs, files in os.walk(folder_path, topdown=False):
            file_count += len(files) - files.count(".DS_Store")
        return file_count


class renameUI(QtWidgets.QWidget, rename_ui.Ui_Form):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setWindowFlag(QtCore.Qt.WindowStaysOnTopHint)
        self.lineEdit__file_path.textChanged.connect(self.updateFolderList)
        self.lineEdit__excel_path.textChanged.connect(self.updateExcelList)
        self.work_thread = WorkThread(
            file_path=pathlib.Path(self.lineEdit__file_path.text()),
            excel_path=pathlib.Path(self.lineEdit__excel_path.text()),
        )
        self.work_thread.start_number = 1001
        self.work_thread.setDaemon = True

        self.work_thread.signals.sig_done.connect(self.rename_finish)
        self.work_thread.signals.sig_update.connect(self.slot_update_progress)
        self.toolButton__file.clicked.connect(self.loadfile)
        self.toolButton__excel.clicked.connect(self.loadfile)
        self.pushButton__start_stop.clicked.connect(self.slot_start)

    def slot_start(self):
        if not self.work_thread.isRunning():
            self.work_thread.start()
            self.pushButton__start_stop.setText("Stop")
        else:
            self.work_thread.set_stop(True)
            self.pushButton__start_stop.setText("Start")

    @QtCore.Slot(int)
    def slot_update_progress(self, val):
        self.progressBar_inpro.setValue(val)

    @QtCore.Slot()
    def rename_finish(self):
        QtWidgets.QMessageBox.information(self, "작업 완료", "모든 파일 및 폴더의 이름 변경 완료!")

    def updateFolderList(self):
        folder_path = self.lineEdit__file_path.text()
        if not folder_path:
            return

        subdirs = [f.name for f in pathlib.Path(folder_path).iterdir() if f.is_dir()]
        self.textBrowser_org.setText("\n".join(subdirs))
        self.updateExcelList()

    def updateExcelList(self):
        excel_path = self.lineEdit__excel_path.text()
        if not excel_path:
            return

        try:
            workbook = load_workbook(filename=excel_path)
            sheet = workbook.active
            data = {}
            for row in sheet.iter_rows(min_row=1, values_only=True):
                if row[0] and row[1]:
                    data[row[0]] = row[1]

            folder_path = self.lineEdit__file_path.text()
            subdirs = [f.name for f in pathlib.Path(folder_path).iterdir() if f.is_dir()]

            new_names = [f" {data.get(subdir, '파일을 어서 넣으세요 그럼 목록이 뜰겁니다.')}" for subdir in subdirs]
            self.textBrowser_after.setText("\n".join(new_names))
        except Exception as e:
            print(f"Error reading excel file: {e}")
            self.textBrowser_after.setText("엑셀 파일 읽기 오류")
            self.updateFolderList()

    def loadfile(self):
        btn: QtWidgets.QToolButton = self.sender()
        if btn.objectName() == "toolButton__file":
            folder_path = QtWidgets.QFileDialog.getExistingDirectory(self, "폴더를 선택하세요")
            if folder_path:
                self.lineEdit__file_path.setText(folder_path)
                self.work_thread.lineEdit__file_path = pathlib.Path(folder_path)
                self.updateFolderList()
        elif btn.objectName() == "toolButton__excel":
            excel_path, _ = QtWidgets.QFileDialog.getOpenFileName(
                self, "엑셀 파일을 선택하세요", "", "Excel Files (*.xlsx *.xls);;All Files (*)"
            )
            if excel_path:
                self.lineEdit__excel_path.setText(excel_path)
                self.work_thread.lineEdit__excel_path = pathlib.Path(excel_path)
                self.updateExcelList()


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = renameUI()
    window.show()
    app.exec()
