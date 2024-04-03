from PIL import Image
import sys
import logging
import os
import importlib
import time
import typing
import subprocess
import pathlib
from PySide6 import QtWidgets, QtCore
from ui import thumb_excel_ui
from libs.qt import qt_lib
from libs.system import sys_library as sys_lib
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as ExcelImage

importlib.reload(qt_lib)
importlib.reload(sys_lib)


class MessageSig:
    message = ""
    is_err = False


class Signals(QtCore.QObject):
    progress_update = QtCore.Signal(int)
    message = QtCore.Signal(MessageSig)
    thumb_path = QtCore.Signal(list)


class ShotListGenerator(QtCore.QThread):
    def __init__(self, startdir: str, targetdir: str):
        super().__init__()
        self.signals = Signals()
        self.__shot_folder_path = startdir
        self.__excel_save_path = targetdir

        self.__is_stop = False

    def set_shot_folder_path(self, val):
        self.__shot_folder_path = val

    def set_excel_save_path(self, val):
        self.__excel_save_path = val

    @property
    def is_stop(self):
        return self.__is_stop

    @is_stop.setter
    def is_stop(self, flag: bool):
        self.__is_stop = flag

    def run(self):
        wb = Workbook()
        ws = wb.active
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center")
        ws["A2"] = "shot_name"
        ws["A2"].font = header_font
        ws["B2"] = "thumbnail"
        ws["B2"].alignment = center_alignment
        ws["B2"].font = header_font
        thumb_lst = list()

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.alignment = center_alignment

        start_row_index = 2

        video_files = [file for file in os.listdir(self.__shot_folder_path) if file.endswith((".mp4", ".avi", ".mov"))]

        for i, filename in enumerate(video_files, start=3):
            ratio = int(i / len(video_files)) * 100
            dst_file = pathlib.Path(self.__excel_save_path) / pathlib.Path(filename).name
            msg_sig = MessageSig()

            try:
                if filename.endswith((".mp4", ".avi", ".mov")):
                    file_path = os.path.join(self.__shot_folder_path, filename)

                    # 엑셀에 파일명 삽입
                    ws[f"A{i}"] = os.path.splitext(filename)[0]

                    # 썸네일 생성
                    thumb_path = os.path.join(self.__shot_folder_path, f"thumb_{filename}.jpg")
                    command = f'ffmpeg -i "{file_path}" -ss 00:00:01.000 -vframes 1 "{thumb_path}"'
                    subprocess.run(command, shell=True)

                    ws.column_dimensions["B"].width = 40
                    ws.row_dimensions[i].height = 130

                    with Image.open(thumb_path) as img:
                        new_width = 270
                        new_height = 150
                        img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
                        img.save(thumb_path)

                    img = ExcelImage(thumb_path)

                    ws.add_image(img, f"B{i}")
                    wb.save(os.path.join(self.__excel_save_path, "shot_list.xlsx"))
                    print(f'Excel 파일이 저장되었습니다: {os.path.join(self.__excel_save_path, "shot_list.xlsx")}')

                    thumb_lst.append(thumb_path)

                    # print(thumb_path, type(thumb_path))
                    # os.remove(thumb_path)
                ################ os.remove(thumb_path)아 -_- 짜증나네 .///

            except FileExistsError as err:
                msg_sig.message = f"{dst_file.as_posix()}"
                msg_sig.is_err = True
                continue
            msg_sig.message = f"[{ratio}%] {filename} -> {dst_file.as_posix()}"
            msg_sig.is_err = False

            if self.__is_stop:
                while True:
                    time.sleep(0.2)
                    print(self.__is_stop)
                    if not self.__is_stop:
                        break
            self.signals.progress_update.emit(ratio)
            self.signals.message.emit(msg_sig)

        self.signals.thumb_path.emit(thumb_lst)


class CustomExcel(QtWidgets.QWidget, thumb_excel_ui.Ui_Form):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setWindowFlag(QtCore.Qt.WindowStaysOnTopHint)
        self.textBrowser__debug__thumb.setReadOnly(True)
        self.__handler = qt_lib.LogHandler(out_stream=self.textBrowser__debug__thumb)
        self.__ui_thread = ShotListGenerator("", "")
        self.toolButton__thum_start.setText("start_path")
        self.toolButton__thum_end.setText("End_path")
        self.toolButton__thum_start.clicked.connect(self.slot_source_dir)
        self.toolButton__thum_end.clicked.connect(self.slot_source_dir)

        self.pushButton_thum_start_stop.clicked.connect(self.toggle_start_stop)
        #self.__ui_thread.signals.progress_update.connect(self.slot_update_progress)
        self.__ui_thread.signals.message.connect(self.slot_print_message)
        self.__ui_thread.signals.thumb_path.connect(self.slot_del_thumbfile)
        self.state = 0b01

    def slot_del_thumbfile(self, thumb_lst):
        if not len(thumb_lst):
            return
        for i in thumb_lst:
            if not os.path.exists(i):
                continue
            os.remove(i)

    @QtCore.Slot(MessageSig)
    def slot_print_message(self, msg: MessageSig):
        if msg.is_err:
            self.__handler.log_msg(logging.error, msg.message)
        else:
            self.__handler.log_msg(logging.info, msg.message)

    # @QtCore.Slot(int)
    # def slot_update_progress(self, val):
    #     self.progressBar_thum.setValue(val)

    def get_all_files(self) -> typing.List[str]:
        dpath = self.textEdit_thum_start.toPlainText()
        return list(sys_lib.System.get_files_recursion(dpath, ["*"]))

    def is_exists_target_dir(self):
        return (
            QtCore.QDir(self.textEdit_thum_end.toPlainText()).exists() and len(self.textEdit_thum_end.toPlainText()) > 0
        )

    def toggle_start_stop(self):
        if self.state & 0b01:
            self.state = 0b10
            self.pushButton_thum_start_stop.setText("Start")  # 버튼 텍스트를 Start로 변경

        else:  # 현재 stop 상태인 경우
            self.state = 0b01  # start 상태로 변경
            self.pushButton_thum_start_stop.setText("Stop")  # 버튼 텍스트를 Stop으로 변경

        if not self.is_exists_target_dir():
            self.__handler.log_msg(logging.error, '엑셀 저장 경로 설정 해주세요 "3" ')
            return

        all_files = self.get_all_files()
        self.__ui_thread.all_files = all_files
        self.__ui_thread.set_excel_save_path(self.textEdit_thum_end.toPlainText())

        if not len(all_files):
            self.handler.log_msg(logging.error, "파일이 없어요 ㅠ_ㅠ")
            return

        self.__ui_thread.set_shot_folder_path(self.textEdit_thum_start.toPlainText())
        self.__ui_thread.set_excel_save_path(self.textEdit_thum_end.toPlainText())
        self.__ui_thread.start()

        self.__ui_thread.daemon = True  # 이 작업은 메인 프로그램 종료시 자동 종료

    def slot_source_dir(self):
        btn: QtWidgets.QToolButton = self.sender()
        sel_dir: pathlib.Path = qt_lib.QtLibs.dir_dialog("/Users/ijiyeong/Desktop/1_program/mov")

        if sel_dir is not None:
            if btn.text() == "start_path":
                self.textEdit_thum_start.setText(sel_dir.as_posix())
            else:
                self.textEdit_thum_end.setText(sel_dir.as_posix())


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    exth = CustomExcel()
    exth.show()
    sys.exit(app.exec_())
